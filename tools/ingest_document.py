"""Registers a document and fills its pages.

`register_document` upserts `source_document` (and, for xlsx, the
spreadsheet_sheet / spreadsheet_cell tables -- formulas kept, because that is
what makes a workbook a template rather than a screenshot).

`extract_pages` fills `source_page` and `source_asset`, and renders each page
to an image under `.tmp/pages/<document_id>/` for a later upload stage.

Two things this corpus breaks a naive implementation on, both handled here:

- **Spread pagination.** Some volumes render two printed pages per PDF page.
  Detected per-page from the pair of page-number words in the footer, not
  from a document-level flag, so it degrades gracefully on the pages (covers)
  that are not spreads even inside a spread-paginated document.
- **Placeholder content.** Lorem-ipsum filler, `TEMPLATE ONLY` stamps and
  `WIP` stamps must be flagged, never ingested as fact. Lorem is detected by
  the classic Latin filler-text vocabulary (the words InDesign's own
  "fill with placeholder text" draws from), not by a single phrase --
  the filler used in this corpus is a scrambled variant that never contains
  the words "lorem" or "ipsum" at all.
"""
from __future__ import annotations

import hashlib
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
import pymupdf
import psycopg

from tools import db

# ── placeholder detection ────────────────────────────────────────────────

# The vocabulary shared by the classic 69-word Lorem Ipsum paragraph and the
# longer "De Finibus" source text that generators (including InDesign's
# built-in filler) draw random sentences from. A page is flagged on the
# *fraction* of its words drawn from this set, not on any single phrase --
# this corpus's filler text never contains "lorem" or "ipsum" itself.
_LATIN_STOPWORDS = frozenset("""
a ab ad adipisci adipiscing alias aliquam aliquid aliquip amet animi
aperiam architecto asperiores aspernatur assumenda at atque aut autem
beatae blanditiis cillum commodi commodo consectetur consequatur consequat
consequuntur corporis corrupti culpa cum cumque cupiditate cupidatat
debitis delectus deleniti deserunt dicta dignissimos distinctio
do dolor dolore dolorem doloremque dolores doloribus dolorum duis
ducimus ea eaque earum eius eiusmod eligendi elit enim eos
error esse est et eu eum eveniet ex excepteur excepturi exercitation
exercitationem expedita explicabo facere facilis fuga fugiat fugit harum
hic id illo illum impedit in incididunt incidunt inventore ipsa ipsam
ipsum irure iste itaque iure iusto labore laboris laboriosam laborum
laudantium libero lorem magna magnam magni maiores maxime minim minima
minus modi molestiae molestias mollit mollitia natus necessitatibus nemo
neque nesciunt nihil nisi nobis non nostrud nostrum nulla numquam
occaecat occaecati odio odit officia officiis omnis optio pariatur
perferendis perspiciatis placeat porro possimus praesentium proident
provident quae quaerat quam quas quasi qui quia quibusdam quidem
quis quisquam quo quod quos ratione recusandae reiciendis rem repellat
repellendus reprehenderit repudiandae rerum saepe sapiente sed sequi
similique sint sit soluta sunt suscipit tempor tempora tempore
temporibus tenetur totam ullam ullamco unde ut vel velit veniam
veritatis vero vitae voluptas voluptate voluptatem voluptates voluptatibus
voluptatum
""".split())

_LOREM_MIN_WORDS = 150
_LOREM_FRACTION = 0.12

_TEMPLATE_ONLY_RE = re.compile(r"template\s*\n?\s*only", re.IGNORECASE)
_WIP_RE = re.compile(r"(?<!\w)WIP(?!\w)")
_WORD_RE = re.compile(r"[a-z]+")

# a document is 'mixed' once placeholder pages stop being a rounding error
_MIXED_DOCUMENT_THRESHOLD = 0.15


def _page_content_status(text: str) -> str:
    """wip > template > lorem > real -- an explicit stamp beats a fuzzy match."""
    if _WIP_RE.search(text):
        return "wip"
    if _TEMPLATE_ONLY_RE.search(text):
        return "template"
    words = _WORD_RE.findall(text.lower())
    if len(words) >= _LOREM_MIN_WORDS:
        hits = sum(1 for w in words if w in _LATIN_STOPWORDS)
        if hits / len(words) > _LOREM_FRACTION:
            return "lorem"
    return "real"


# ── spread pagination ────────────────────────────────────────────────────

_PAGE_NUM_RE = re.compile(r"\A\d{1,4}\Z")


def _printed_page_label(page: "pymupdf.Page") -> str | None:
    """'187 / 188' when this page is a landscape spread of two printed pages.

    Detected from the pair of standalone page-number words sitting in the
    bottom margin, one in each half of the page -- not from a document-level
    flag, so a cover page inside an otherwise spread-paginated volume is
    correctly left alone.
    """
    rect = page.rect
    if rect.height == 0 or (rect.width / rect.height) < 1.3:
        return None
    band = rect.height * 0.08  # bottom margin band
    half = rect.width / 2
    left_nums: list[str] = []
    right_nums: list[str] = []
    try:
        words = page.get_text("words")
    except Exception:
        return None
    for x0, y0, x1, y1, text, *_ in words:
        if y0 < rect.height - band:
            continue
        if not _PAGE_NUM_RE.match(text):
            continue
        (left_nums if x0 < half else right_nums).append(text)
    if len(left_nums) == 1 and len(right_nums) == 1:
        try:
            a, b = int(left_nums[0]), int(right_nums[0])
        except ValueError:
            return None
        lo, hi = min(a, b), max(a, b)
        return f"{lo} / {hi}"
    return None


# ── page image rendering ─────────────────────────────────────────────────

_TARGET_WIDTH_PX = 1400
_WEBP_QUALITY = 75
_PAGES_DIR = Path(".tmp") / "pages"


def _object_key(out_path: Path) -> str:
    """Local render path -> the object key stored in the database.

    One definition, matching tools/upload_page_images.py's _key_for(), because
    a key that resolves on one database and not the other is a broken image
    that only appears in production.
    """
    return "pages/" + out_path.relative_to(_PAGES_DIR).as_posix()


def _render_page_image(page: "pymupdf.Page", out_dir: Path, page_index: int) -> str | None:
    """Render to WebP (~1400px wide) under out_dir. Returns the OBJECT KEY.

    The key, not the path. It used to return `str(out_path)` -- the local
    `.tmp/pages/<uuid>/00001.webp` -- while every consumer expects the R2 key
    `pages/<uuid>/00001.webp`: tools/upload_page_images.py strips the prefix at
    upload time, the web route matches source_page.page_image_key verbatim, and
    workflows/deploy_web.md states the convention outright. All 806 live rows
    hold the key form, so the mismatch never showed; the next document ingested
    would simply have had an unresolvable image key, on both databases, and the
    only symptom is a missing page scan.
    """
    rect = page.rect
    if rect.width <= 0:
        return None
    zoom = max(0.3, min(4.0, _TARGET_WIDTH_PX / rect.width))
    try:
        pix = page.get_pixmap(matrix=pymupdf.Matrix(zoom, zoom))
    except Exception:
        return None

    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{page_index:05d}"
    cwebp = shutil.which("cwebp")
    if cwebp:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            pix.save(tmp_path)
            out_path = out_dir / f"{stem}.webp"
            result = subprocess.run(
                [cwebp, "-quiet", "-q", str(_WEBP_QUALITY), str(tmp_path), "-o", str(out_path)],
                capture_output=True,
            )
            if result.returncode == 0 and out_path.exists():
                return _object_key(out_path)
        except Exception:
            pass
        finally:
            tmp_path.unlink(missing_ok=True)

    # cwebp unavailable or failed -- fall back to PNG rather than lose the page
    out_path = out_dir / f"{stem}.png"
    try:
        pix.save(out_path)
        return _object_key(out_path)
    except Exception:
        return None


# ── source_document ──────────────────────────────────────────────────────

_DOCUMENT_COLUMNS = {
    "title", "series_ref", "revision", "version_label", "issue_date",
    "client_org_id", "author_org_id", "consultant_org_ids", "confidentiality",
    "content_status", "language", "is_spread_paginated", "original_filename",
}


def register_document(conn: psycopg.Connection, *, path: Path, sha256: str, slug: str,
                       doc_kind: str, meta: dict[str, Any]) -> str:
    """Upsert source_document (+ spreadsheet_sheet/cell for xlsx). Returns document_id."""
    path = Path(path)
    fields = {k: v for k, v in meta.items() if k in _DOCUMENT_COLUMNS}
    fields.setdefault("original_filename", path.name)
    size_bytes = path.stat().st_size if path.exists() else meta.get("size_bytes")

    existing = db.one(conn, "SELECT id FROM source_document WHERE sha256 = %s", (sha256,))
    if existing:
        document_id = existing["id"]
        # A re-ingest can legitimately relabel a document (reclassification,
        # a corrected slug); retire whatever OTHER row currently holds the
        # target slug first, so this UPDATE can't collide with the partial
        # unique index on (slug) WHERE is_current.
        conn.execute(
            "UPDATE source_document SET is_current = false WHERE slug = %s AND is_current AND id <> %s",
            (slug, document_id),
        )
        extra_cols = "".join(f", {c} = %s" for c in fields)
        conn.execute(
            f"""UPDATE source_document
                   SET doc_kind = %s, slug = %s, size_bytes = %s, is_current = true {extra_cols}
                 WHERE id = %s""",
            [doc_kind, slug, size_bytes, *fields.values(), document_id],
        )
    else:
        prev = db.one(
            conn,
            "SELECT id FROM source_document WHERE slug = %s AND is_current",
            (slug,),
        )
        insert_values = dict(fields)
        insert_values.update(slug=slug, doc_kind=doc_kind, sha256=sha256, size_bytes=size_bytes)
        if prev:
            # Retire the old revision BEFORE inserting the new one. The partial
            # unique index on (slug) WHERE is_current is not deferrable, so both
            # rows carrying is_current for an instant is a constraint violation,
            # not a transient state. Same transaction either way.
            insert_values["supersedes_id"] = prev["id"]
            conn.execute(
                "UPDATE source_document SET is_current = false WHERE id = %s",
                (prev["id"],),
            )
        document_id = db.insert_returning_id(conn, "source_document", insert_values)

    if path.suffix.lower() == ".xlsx":
        _register_spreadsheet(conn, document_id, path)

    return str(document_id)


def _register_spreadsheet(conn: psycopg.Connection, document_id: str, path: Path) -> None:
    # openpyxl cannot hand back both the formula text and its cached value
    # from one load, so the workbook is opened twice.
    wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # pass 1: every cell any formula in the workbook refers to, so role
    # classification (a referenced constant is an 'input', not a stray label)
    # can see across sheets.
    referenced: set[tuple[str, str]] = set()
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    for ref_sheet, ref_coord in _formula_refs(cell.value, sheet_name):
                        referenced.add((ref_sheet, ref_coord))

    for ordinal, sheet_name in enumerate(wb_formulas.sheetnames):
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else None
        sheet_row = db.one(
            conn,
            """INSERT INTO spreadsheet_sheet (document_id, name, ordinal)
                    VALUES (%s, %s, %s)
               ON CONFLICT (document_id, name) DO UPDATE SET ordinal = EXCLUDED.ordinal
               RETURNING id""",
            (document_id, sheet_name, ordinal),
        )
        sheet_id = sheet_row["id"]

        for row in ws_f.iter_rows():
            for cell in row:
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                raw_value = cell.value if formula is None else None
                cached_value = None
                if ws_v is not None:
                    try:
                        cached_value = ws_v[cell.coordinate].value
                    except Exception:
                        cached_value = None

                if formula is None and raw_value is None:
                    continue  # empty cell -- nothing to store

                value_numeric = None
                value_text = None
                effective = cached_value if formula else raw_value
                if isinstance(effective, bool):
                    value_text = str(effective)
                elif isinstance(effective, (int, float)):
                    value_numeric = effective
                elif effective is not None:
                    value_text = str(effective)

                role = _cell_role(
                    formula=formula,
                    is_numeric=value_numeric is not None,
                    is_text=value_text is not None,
                    referenced=(sheet_name, cell.coordinate) in referenced,
                )

                conn.execute(
                    """INSERT INTO spreadsheet_cell
                            (sheet_id, ref, row_num, col_num, value_text, value_numeric,
                             formula, number_format, role)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       ON CONFLICT (sheet_id, ref) DO UPDATE SET
                            row_num = EXCLUDED.row_num, col_num = EXCLUDED.col_num,
                            value_text = EXCLUDED.value_text, value_numeric = EXCLUDED.value_numeric,
                            formula = EXCLUDED.formula, number_format = EXCLUDED.number_format,
                            role = EXCLUDED.role""",
                    (sheet_id, cell.coordinate, cell.row, cell.column, value_text,
                     value_numeric, formula, cell.number_format, role),
                )

    wb_formulas.close()
    wb_values.close()


_CELL_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_]+))?!?\$?([A-Z]{1,3})\$?(\d{1,7})"
)


def _formula_refs(formula: str, home_sheet: str) -> list[tuple[str, str]]:
    """Best-effort (sheet, coordinate) pairs a formula string mentions."""
    out = []
    for quoted, bare, col, row in _CELL_REF_RE.findall(formula):
        sheet = quoted or bare or home_sheet
        out.append((sheet, f"{col}{row}"))
    return out


def _cell_role(*, formula: str | None, is_numeric: bool, is_text: bool, referenced: bool) -> str | None:
    if formula is not None:
        return "calc" if referenced else "output"
    if is_numeric:
        return "input"
    if is_text:
        return "label"
    return None


# ── source_page / source_asset ───────────────────────────────────────────


def extract_pages(conn: psycopg.Connection, document_id: str, path: Path) -> int:
    """Fill source_page + source_asset, render page images. Returns page count."""
    path = Path(path)
    if path.suffix.lower() != ".pdf":
        return 0  # xlsx has no pages; its content lives in spreadsheet_cell

    out_dir = _PAGES_DIR / str(document_id)
    placeholder_count = 0
    total = 0

    with pymupdf.open(path) as doc:
        total = doc.page_count
        for i in range(total):
            page = doc[i]
            page_index = i + 1
            try:
                text = page.get_text()
            except Exception:
                text = ""
            text = text.replace("\x00", "")  # postgres text cannot hold NUL bytes
            content_status = _page_content_status(text)
            if content_status != "real":
                placeholder_count += 1

            try:
                image_count = len(page.get_images(full=True))
            except Exception:
                image_count = 0
            try:
                vector_op_count = len(page.get_drawings())
            except Exception:
                vector_op_count = 0

            printed_page_label = _printed_page_label(page)
            image_key = _render_page_image(page, out_dir, page_index)

            page_row = db.one(
                conn,
                """INSERT INTO source_page
                        (document_id, page_index, printed_page_label, width_pt, height_pt,
                         text, image_count, vector_op_count, content_status, page_image_key)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (document_id, page_index) DO UPDATE SET
                        printed_page_label = EXCLUDED.printed_page_label,
                        width_pt = EXCLUDED.width_pt, height_pt = EXCLUDED.height_pt,
                        text = EXCLUDED.text, image_count = EXCLUDED.image_count,
                        vector_op_count = EXCLUDED.vector_op_count,
                        content_status = EXCLUDED.content_status,
                        page_image_key = EXCLUDED.page_image_key
                   RETURNING id""",
                (document_id, page_index, printed_page_label, page.rect.width, page.rect.height,
                 text, image_count, vector_op_count, content_status, image_key),
            )
            page_id = page_row["id"]

            # Upsert on (page_id, sha256) rather than the delete-then-insert
            # this used to be. The blanket DELETE was correct while nothing but
            # this function wrote to source_asset, and destroys a figure
            # description the moment something does -- silently, re-creating the
            # row empty beside it. The hash of the decoded bytes is the asset's
            # identity: same image, same row, across extraction runs.
            seen: list[str] = []
            try:
                images = page.get_images(full=True)
            except Exception:
                images = []
            for img in images:
                xref = img[0]
                try:
                    rects = page.get_image_rects(xref)
                except Exception:
                    rects = []
                # Only the first placement is recorded. A graphic drawn several
                # times on one page (a repeated logo) is one row, and a NULL
                # bbox is a real state: get_image_rects comes back empty for
                # pattern and SMask-only xrefs, and the row is still worth
                # keeping because the image is still on the page.
                bbox = [round(v, 2) for v in rects[0]] if rects else None
                try:
                    sha = hashlib.sha256(doc.extract_image(xref)["image"]).hexdigest()
                except Exception:
                    sha = None
                if sha is not None:
                    seen.append(sha)
                # Adopt a pre-sha256 row for the same box before inserting.
                # Without this the guard below preserves the old row and the
                # insert adds a hashed twin beside it, so every described asset
                # ends up duplicated on the first re-ingest. Matching on bbox is
                # sound here because a row's box is exactly what identified it
                # before the hash existed.
                adopted = conn.execute(
                    "UPDATE source_asset SET sha256 = %s WHERE page_id = %s "
                    # The cast is required: bbox is numeric(9,2)[] and psycopg
                    # sends a Python float list as double precision[], for which
                    # Postgres has no equality operator against numeric[].
                    "AND sha256 IS NULL AND bbox IS NOT DISTINCT FROM %s::numeric(9,2)[]",
                    (sha, page_id, bbox),
                ).rowcount if sha is not None else 0
                if not adopted:
                    conn.execute(
                        """INSERT INTO source_asset (page_id, sha256, bbox)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (page_id, sha256) WHERE sha256 IS NOT NULL
                           DO UPDATE SET bbox = EXCLUDED.bbox""",
                        (page_id, sha, bbox),
                    )
            # Anything left on this page that the document no longer contains.
            #
            # Two guards, not one. The hash match handles rows written since
            # sha256 existed. The `vlm_description IS NULL AND caption IS NULL`
            # clause handles everything older: those rows cannot be matched by
            # hash, so without it the first re-ingest after a describe run would
            # delete exactly the assets someone had spent money describing. A
            # stale row for a figure genuinely removed from the document is a
            # visible, fixable wart; silently destroying generated content is
            # neither. Nothing this stage did not write is this stage's to
            # delete.
            conn.execute(
                "DELETE FROM source_asset WHERE page_id = %s "
                "AND (sha256 IS NULL OR NOT (sha256 = ANY(%s))) "
                "AND vlm_description IS NULL AND caption IS NULL",
                (page_id, seen),
            )

    conn.execute(
        "UPDATE source_document SET page_count = %s WHERE id = %s",
        (total, document_id),
    )
    if total and placeholder_count / total > _MIXED_DOCUMENT_THRESHOLD:
        conn.execute(
            "UPDATE source_document SET content_status = 'mixed' WHERE id = %s",
            (document_id,),
        )
    # Roll the per-page evidence up to the document. A quarter is enough: these
    # volumes carry covers, dividers and Word-origin appendices that are single
    # pages inside an otherwise spread-paginated book, and the flag exists to
    # tell a citation renderer that page_index and the printed number differ.
    conn.execute(
        """UPDATE source_document d SET is_spread_paginated = (
               SELECT count(p.printed_page_label) > 0.25 * count(*)
               FROM source_page p WHERE p.document_id = d.id)
           WHERE d.id = %s""",
        (document_id,),
    )

    return total
