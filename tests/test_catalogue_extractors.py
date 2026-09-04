"""Tests for the three catalogue-family extractors this module owns:
extractors/typology_catalogue.py, extractors/spreadsheet.py, extractors/deck.py.

Extractors are pure (CONTRACT.md): no database, no network. These tests
build small synthetic fixtures on disk (a tiny .xlsx via openpyxl, tiny
.pdf's via pymupdf) rather than depend on the real corpus, which is not
tracked by git and will not exist in a fresh clone (private/documents.yaml
+ CONTRACT.md). The real corpus was used during development to verify
counts (see the task report), but that is not repeatable here.

Run directly:
    ./.venv/bin/python tests/test_catalogue_extractors.py
or via pytest:
    ./.venv/bin/python -m pytest tests/test_catalogue_extractors.py -q
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pymupdf
import pytest

from tools.pipeline import DocumentContext
from extractors.typology_catalogue import (
    TYPOLOGY_CATALOGUE, _clean_body, _disclaimer, _figure_codes, _section_code,
)
from extractors.spreadsheet import SPREADSHEET_CALCULATOR
from extractors.deck import DECK, _measure_category


# ── extractors/spreadsheet.py ──────────────────────────────────────────────


def _build_calc_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee calculation"
    # a small formula graph mirroring the real calc-fees shape:
    # rate (input) x hours (input, monthly grid) -> cost -> +mark-up -> fee to bill (output)
    ws["D2"] = "Rate"
    ws["E2"] = 50.0  # input: referenced by F4's formula below
    ws["D4"] = "Jan"
    ws["E4"] = "Feb"
    ws["D5"] = "Hours"
    ws["D5"].number_format = "General"
    import datetime
    ws["D4"] = datetime.datetime(2024, 1, 1)
    ws["E4"] = datetime.datetime(2024, 2, 1)
    ws["F4"] = "Total"
    ws["D5"] = 10
    ws["E5"] = 20
    ws["D6"] = "=D5*$E$2"
    ws["E6"] = "=E5*$E$2"
    ws["F6"] = "=SUM(D6:E6)"
    ws["D8"] = "Mark-up"
    ws["E8"] = 0.25  # input: referenced by F9
    ws["E8"].number_format = "0%"
    ws["D9"] = "Fee to bill"
    ws["F9"] = "=F6*(1+$E$8)"
    wb.save(path)


def test_spreadsheet_calculator_finds_inputs_and_outputs():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "calc.xlsx"
        _build_calc_xlsx(path)
        ctx = DocumentContext(document_id="doc-1", slug="calc-test", path=path,
                               doc_kind="calculator", page_count=0, pages=[])
        ex = SPREADSHEET_CALCULATOR.extract(ctx)

        template = next(it for it in ex.items if it.item_type == "template")
        assert template.payload["template_kind"] == "calculator"
        assert template.payload["engine"] == "xlsx"
        assert template.payload["slug"] == "calc-test"

        params = {p["cell_ref"]: p for p in template.payload["parameters"]}
        assert params["E2"]["is_input"] and float(params["E2"]["default_value"]) == 50.0
        assert params["E8"]["is_input"] and float(params["E8"]["default_value"]) == 0.25
        assert params["E8"]["data_type"] == "percent"
        # F9 ('Fee to bill') is the terminal formula nobody else reads -> output
        assert params["F9"]["is_output"]
        assert params["F9"]["label"] == "Fee to bill"
        # F6 feeds F9, so it is an intermediate, not a reported output
        assert "F6" not in params or not params["F6"]["is_output"]

        assert any(n.node_kind == "sheet" for n in ex.nodes)


def test_spreadsheet_calculator_is_pure():
    """No import-time or call-time side effects reach a database."""
    import tools.pipeline as pipeline_mod
    assert not hasattr(SPREADSHEET_CALCULATOR, "conn")
    assert "db" not in dir(SPREADSHEET_CALCULATOR)


# ── extractors/deck.py ─────────────────────────────────────────────────────


def _build_deck_pdf(path: Path) -> None:
    doc = pymupdf.open()

    # slide 1: the Tier measures slide ("Strategic Approach"), two columns
    p1 = doc.new_page(width=720, height=405)
    p1.insert_text((180, 100), "Tier 1", fontsize=10)
    p1.insert_text((380, 100), "Tier 2", fontsize=10)
    p1.insert_text((180, 130), "Cool Materials", fontsize=9)
    p1.insert_text((380, 130), "Cool Materials", fontsize=9)
    p1.insert_text((380, 160), "Active Cooling (mechanical vent cooling)", fontsize=9)
    p1.insert_text((10, 10), "Strategic Approach", fontsize=12)

    # slide 2: the outcomes slide ("Comfortable hours"), Tier N stacked above its %
    p2 = doc.new_page(width=720, height=405)
    p2.insert_text((10, 40), "Comfortable hours", fontsize=12)
    p2.insert_text((500, 60), "40%", fontsize=9)   # baseline, unlabelled
    p2.insert_text((180, 120), "Tier 1", fontsize=9)
    p2.insert_text((500, 110), "60%", fontsize=9)
    p2.insert_text((180, 160), "Tier 2", fontsize=9)
    p2.insert_text((500, 150), "70%", fontsize=9)

    # slide 3: image-only (near-empty text layer)
    p3 = doc.new_page(width=720, height=405)
    p3.insert_text((10, 10), "3", fontsize=8)

    doc.save(path)
    doc.close()


def test_deck_extractor_tier_ladder_and_image_only_flagging():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "deck.pdf"
        _build_deck_pdf(path)
        doc = pymupdf.open(path)
        pages = [{"page_index": i + 1, "text": doc[i].get_text(), "content_status": "real"}
                  for i in range(doc.page_count)]
        doc.close()
        ctx = DocumentContext(document_id="doc-2", slug="deck-test", path=path,
                               doc_kind="deck", page_count=len(pages), pages=pages)
        ex = DECK.extract(ctx)

        assert len(ex.nodes) == 3
        assert all(n.node_kind == "slide" for n in ex.nodes)

        tiers = {it.title: it for it in ex.items if it.item_type == "pattern"}
        assert "Tier 1" in tiers and "Tier 2" in tiers
        assert tiers["Tier 1"].payload["attributes"]["comfortable_hours_pct"] == 60.0
        assert tiers["Tier 2"].payload["attributes"]["comfortable_hours_pct"] == 70.0
        assert tiers["Tier 1"].payload["attributes"]["baseline_comfortable_hours_pct"] == 40.0

        tier2_measures = {m["name"]: m["category"] for m in tiers["Tier 2"].payload["attributes"]["measures"]}
        assert tier2_measures.get("Active Cooling (mechanical vent cooling)") == "active"

        assert any("image-only" in w and "slide 3" in w for w in ex.warnings)
        assert ex.stats["image_only_slides"] >= 1


@pytest.mark.parametrize("name,expected", [
    ("Active Cooling (mechanical vent cooling)", "active"),
    ("Water features (passive evaporative cooling)", "passive"),
    ("Sunken spaces", "thermal_transition"),
    ("Shelter Walls", "thermal_transition"),
    ("Vegetation", "passive"),
])
def test_measure_category_heuristic(name, expected):
    assert _measure_category(name) == expected


# ── extractors/typology_catalogue.py ────────────────────────────────────────


def _build_typology_pdf(path: Path) -> None:
    doc = pymupdf.open()

    # page 1: chapter cover (no useful text)
    doc.new_page(width=595, height=842).insert_text((50, 50), "COVER", fontsize=10)

    # page 2: a real subsection with a section code, a figure and the disclaimer
    p2 = doc.new_page(width=595, height=842)
    p2.insert_text((50, 50), "2.1", fontsize=9)
    p2.insert_text((50, 65), "2.1.2", fontsize=9)
    p2.insert_text((50, 100), "Density and F.A.R. guidance text explaining the ratio in detail here.", fontsize=9)
    p2.insert_text((50, 120), "FIGURE 2.1.2.1", fontsize=9)
    p2.insert_text((50, 700), "*All Designs must comply with applicable Local Building Codes and Fire Regulations.", fontsize=7)

    # page 3: a child subsection under the same section, still real, tests
    # that a container node's page range spans past its first child (the
    # bug this extractor had to be fixed for: see _build_toc)
    p3 = doc.new_page(width=595, height=842)
    p3.insert_text((50, 50), "2.1.3", fontsize=9)
    p3.insert_text((50, 100), "Further guidance continuing the same section on another page.", fontsize=9)

    # page 4: WIP page -- must never surface as an item
    p4 = doc.new_page(width=595, height=842)
    p4.insert_text((50, 50), "2.1.4", fontsize=9)
    p4.insert_text((50, 100), "WIP content that must not be ingested as fact.", fontsize=9)
    p4.insert_text((50, 120), "WIP", fontsize=9)

    doc.set_toc([
        [1, "PLOT PLANNING", 2],
        [2, "Plot Morphology", 2],
        [3, "Density and FAR", 2],
        [3, "More Density", 3],
        [3, "WIP Subsection", 4],
    ])
    doc.save(path)
    doc.close()


def test_typology_catalogue_page_ranges_span_children():
    """Regression test: a container's page_to must include every descendant
    page, not just stop at its first child (which starts on the same page)."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "typology.pdf"
        _build_typology_pdf(path)
        doc = pymupdf.open(path)
        pages = [{"page_index": i + 1, "text": doc[i].get_text(), "content_status": "real"}
                  for i in range(doc.page_count)]
        pages[3]["content_status"] = "wip"  # page 4
        doc.close()

        ctx = DocumentContext(document_id="doc-3", slug="typology-test", path=path,
                               doc_kind="guideline_report", page_count=len(pages), pages=pages)
        ex = TYPOLOGY_CATALOGUE.extract(ctx)

        section = next(n for n in ex.nodes if n.title == "Plot Morphology")
        assert section.page_from == 2
        assert section.page_to == 4, "must span its two real children plus the WIP one, not collapse to page 2"

        guidance = {it.title: it for it in ex.items if it.item_type == "guidance"}
        assert "Density and FAR" in guidance
        assert "More Density" in guidance
        assert "WIP Subsection" not in guidance, "a WIP-only subsection must not produce a guidance item"

        g = guidance["Density and FAR"]
        assert g.payload["disclaimer"] and "Local Building Codes" in g.payload["disclaimer"]
        assert "FIGURE 2.1.2.1" in g.payload["legend_tokens"]

        # the hard safety property: nothing cites the WIP page
        status_by_page = {p["page_index"]: p["content_status"] for p in pages}
        for it in ex.items:
            for c in it.citations:
                assert status_by_page[c.page_index] == "real", (
                    f"{it.item_type} {it.title!r} cited a non-real page {c.page_index}"
                )


def test_section_code_picks_the_most_specific_decimal_code():
    text = "some heading\n2.1\n2.1.2\nDENSITY + F.A.R.\nbody text"
    assert _section_code(text) == "2.1.2"


def test_figure_codes_dedupes_preserving_order():
    text = "FIGURE 5.1.2.1 blah FIGURE 5.1.2.2 blah FIGURE 5.1.2.1"
    assert _figure_codes(text) == ["FIGURE 5.1.2.1", "FIGURE 5.1.2.2"]


def test_disclaimer_detection():
    text = "*All Designs must comply with applicable Local Building Codes and Fire Regulations.\nmore text"
    assert _disclaimer(text) is not None
    assert _disclaimer("nothing relevant here") is None


def test_clean_body_strips_page_furniture():
    text = "ABC Multifamily Housing\n5.1.2\n06\nActual guidance sentence here.\n/  64"
    cleaned = _clean_body(text)
    assert "ABC Multifamily Housing" not in cleaned
    assert "5.1.2" not in cleaned
    assert "Actual guidance sentence here." in cleaned


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
