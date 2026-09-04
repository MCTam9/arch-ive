"""Extractor for xlsx calculators (doc_kind='calculator').

Targets calc-fees, calc-budget-cost-rates, calc-cashflow. Another module
already loads raw cells into spreadsheet_sheet/spreadsheet_cell -- this
module does not touch that. It reads the workbook directly with openpyxl and
emits the *template*: the calculator's formula-graph interface.

The rule (per CONTRACT.md): a cell holding a constant but referenced by some
formula is an INPUT; a formula cell nobody references is an OUTPUT. Every
candidate is labelled from the nearest text cell to its left or above,
because that is how these sheets are laid out.

Two refinements earned by reading the real workbooks:

- Grid collapsing: a row repeated across ~30 monthly columns (one FTE
  allocation or cost line per month) is one parameter, not thirty. We detect
  the month/date header row once per sheet and, for any candidate that falls
  in it, represent the whole row by its "Total"/"Totals" column if the sheet
  has one, else its first populated cell.
- Forced outputs: strict "nobody references it" misses headline figures that
  feed a light downstream transform (a ROUND(), a ratio) -- 'Fee to bill' is
  read by 'Minimum monthly invoice', 'Overhead percentage' is read by every
  grade's hourly rate. Both are still the named deliverables CONTRACT.md
  calls out, so their labels are force-included as outputs alongside the
  strict-terminal set.
"""
from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tools.pipeline import DocumentContext, Extraction, Item, Node, register

# Labels whose formula is a genuine calculator output even though a sibling
# row reads it for a rounded/ratio view immediately below it. Matched
# case-insensitively against the row/cell's own nearest-label text.
_FORCE_OUTPUT_LABELS = {
    "fee to bill",
    "invoice",
    "profit",
    "profit percentage",
    "overhead percentage",
    "total turnover",
    "total cash in booked",
    "net profit",
}

_EXCEL_EPOCH = datetime.date(1899, 12, 30)


def _excel_serial_to_date(value: float) -> datetime.date | None:
    """Best-effort conversion for a bare Excel date serial (openpyxl only
    auto-converts cells that already carry a date number_format)."""
    if not (20000 < value < 60000):  # ~1954-2064; keeps this from firing on ordinary counts
        return None
    try:
        return _EXCEL_EPOCH + datetime.timedelta(days=value)
    except OverflowError:
        return None


def _is_percent_format(number_format: str | None) -> bool:
    return bool(number_format) and "%" in number_format


def _is_currency_format(number_format: str | None) -> bool:
    return bool(number_format) and ("£" in number_format or "$" in number_format or "£" in number_format)


def _is_date_format(number_format: str | None) -> bool:
    if not number_format:
        return False
    nf = number_format.lower()
    return any(tok in nf for tok in ("yy", "mmm", "dd/mm", "mm/dd"))


def _formula_refs(formula: str, default_sheet: str) -> list[tuple[str, str]]:
    """(sheet_name, cell_coord) pairs a formula's tokens resolve to. Ranges are
    expanded (capped) rather than kept as ranges, so the result can be tested
    against a per-sheet set of individual referenced coordinates."""
    out: list[tuple[str, str]] = []
    try:
        tok = Tokenizer(formula)
    except Exception:
        return out
    for item in tok.items:
        if item.type != "OPERAND" or item.subtype != "RANGE":
            continue
        text = item.value
        sheet = default_sheet
        ref = text
        if "!" in text:
            sheet_part, ref = text.rsplit("!", 1)
            sheet = sheet_part.strip("'").replace("''", "'")
        ref = ref.replace("$", "")
        if ":" in ref:
            start, end = ref.split(":", 1)
            out.extend((sheet, c) for c in _expand_range(start, end))
        elif ref and ref[0].isalpha():
            out.append((sheet, ref))
    return out


def _expand_range(start: str, end: str, cap: int = 3000) -> list[str]:
    try:
        c1, r1 = coordinate_from_string(start)
        c2, r2 = coordinate_from_string(end)
        ci1, ci2 = column_index_from_string(c1), column_index_from_string(c2)
    except ValueError:
        return []
    cells: list[str] = []
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for ci in range(min(ci1, ci2), max(ci1, ci2) + 1):
            cells.append(f"{get_column_letter(ci)}{r}")
            if len(cells) >= cap:
                return cells
    return cells


def _collect_references(wb) -> dict[str, set[str]]:
    refs: dict[str, set[str]] = {ws.title: set() for ws in wb.worksheets}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    for sheet_name, coord in _formula_refs(v, ws.title):
                        refs.setdefault(sheet_name, set()).add(coord)
    return refs


def _period_grid(ws: Worksheet) -> tuple[set[str], str | None, int | None]:
    """(date-header columns, 'Total' column letter or None, header row) for a
    sheet whose data rows repeat across monthly/period columns -- true of
    every sheet in these three workbooks."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        date_cols = {c.column_letter for c in row if isinstance(c.value, datetime.datetime)}
        if len(date_cols) >= 3:
            total_col = None
            for c in row:
                if isinstance(c.value, str) and c.value.strip().lower() in ("total", "totals", "year end"):
                    total_col = c.column_letter
            return date_cols, total_col, row[0].row
    return set(), None, None


def _nearest_label(ws: Worksheet, row: int, col_letter: str) -> str | None:
    """Nearest text-only cell to the left in the same row, else nearest above
    in the same column -- how these sheets place their row/column labels."""
    col = column_index_from_string(col_letter)
    best: tuple[int, str] | None = None
    for c in range(col - 1, max(col - 12, 0), -1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and not v.startswith("="):
            dist = col - c
            if best is None or dist < best[0]:
                best = (dist, v.strip())
            break  # nearest non-empty text cell going left is enough
    if best:
        return best[1]
    for r in range(row - 1, max(row - 12, 0), -1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip() and not v.startswith("="):
            return v.strip()
    return None


def _data_type_and_unit(number_format: str | None, sample: Any) -> tuple[str, str | None]:
    if _is_percent_format(number_format):
        return "percent", None
    if _is_currency_format(number_format):
        return "number", "gbp"
    if _is_date_format(number_format) or isinstance(sample, datetime.datetime):
        return "date", None
    return "number", None


class SpreadsheetCalculatorExtractor:
    doc_kinds: tuple[str, ...] = ("calculator",)

    def extract(self, ctx: DocumentContext) -> Extraction:
        ex = Extraction()
        wb_f = openpyxl.load_workbook(ctx.path, data_only=False)
        try:
            wb_v = openpyxl.load_workbook(ctx.path, data_only=True)
        except Exception:
            wb_v = None

        refs = _collect_references(wb_f)
        parameters: list[dict[str, Any]] = []
        ordinal = 0
        seen_rows: set[tuple[str, int, str]] = set()  # (sheet, row, "in"/"out") already represented

        ex.units.append({"id": "gbp", "symbol": "£", "dimension": "currency", "si_factor": None})

        for sheet_idx, ws in enumerate(wb_f.worksheets, start=1):
            ws_v = wb_v[ws.title] if wb_v is not None else None
            sheet_refs = refs.get(ws.title, set())
            grid_cols, total_col, header_row = _period_grid(ws)

            ex.nodes.append(
                Node(node_kind="sheet", title=ws.title, ordinal=sheet_idx, ref=f"sheet-{sheet_idx}")
            )

            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None or cell.row == header_row:
                        continue
                    col_letter = cell.column_letter

                    # A row already represented by its Total column via grid
                    # collapsing (below) must not also be picked up here when
                    # the loop reaches that same Total cell directly.
                    if total_col and col_letter == total_col and (
                        (ws.title, cell.row, "in") in seen_rows or (ws.title, cell.row, "out") in seen_rows
                    ):
                        continue

                    is_formula = isinstance(v, str) and v.startswith("=")
                    is_constant_number = isinstance(v, (int, float)) and not isinstance(v, bool)
                    if not is_formula and not is_constant_number:
                        continue

                    referenced = cell.coordinate in sheet_refs
                    in_grid = col_letter in grid_cols
                    label = _nearest_label(ws, cell.row, col_letter)

                    if is_formula:
                        forced = bool(label) and label.strip().lower() in _FORCE_OUTPUT_LABELS
                        if referenced and not forced:
                            continue  # feeds another formula -- an intermediate, not an output
                        cached = ws_v[cell.coordinate].value if ws_v is not None else None
                        # a definite text cached value (e.g. a formula that just mirrors a
                        # grade name from another sheet) is not a metric; a *missing* cached
                        # value (the workbook was never opened by a real calc engine, as in a
                        # hand-built test fixture) is inconclusive, not disqualifying
                        if isinstance(cached, str) or isinstance(cached, bool):
                            continue
                        kind = "out"
                    elif is_constant_number:
                        if not referenced:
                            continue  # a bare number nobody uses is not a parameter
                        kind = "in"
                    else:
                        continue

                    if in_grid:
                        row_key = (ws.title, cell.row, kind)
                        if row_key in seen_rows:
                            continue
                        seen_rows.add(row_key)
                        last_col = sorted(grid_cols, key=column_index_from_string)[-1]
                        range_ref = f"{sorted(grid_cols, key=column_index_from_string)[0]}{cell.row}:{last_col}{cell.row}"
                        if kind == "out" and total_col:
                            # the row's rollup formula is the meaningful output value
                            rep_cell = ws.cell(row=cell.row, column=column_index_from_string(total_col))
                            target_cell = rep_cell if rep_cell.value is not None else cell
                            cell_ref = target_cell.coordinate
                        else:
                            # an input row (e.g. monthly FTE allocation) has no single
                            # cell to point at -- represent the whole range and
                            # summarise its constants below
                            target_cell = cell
                            cell_ref = range_ref
                    else:
                        target_cell = cell
                        cell_ref = cell.coordinate

                    data_type, unit_id = _data_type_and_unit(target_cell.number_format, target_cell.value)
                    if in_grid and kind == "in":
                        default_value = self._summarise_row(ws, cell.row, grid_cols)
                    else:
                        default_value = self._default_value(target_cell, ws_v)
                    ordinal += 1
                    parameters.append({
                        "name": _slugify(label) or f"{ws.title.lower().replace(' ', '_')}_{target_cell.coordinate.lower()}",
                        "label": label,
                        "sheet_name": ws.title,
                        "cell_ref": cell_ref,
                        "data_type": data_type,
                        "unit_id": unit_id,
                        "default_value": default_value,
                        "is_input": kind == "in",
                        "is_output": kind == "out",
                        "ordinal": ordinal,
                    })

            for name, default in self._inline_constants(ws):
                ordinal += 1
                parameters.append({
                    "name": name,
                    "label": name.replace("_", " ").title(),
                    "sheet_name": ws.title,
                    "cell_ref": None,
                    "data_type": "percent" if name.startswith("probability_weighting_") else "number",
                    "unit_id": None,
                    "default_value": default,
                    "is_input": True,
                    "is_output": False,
                    "ordinal": ordinal,
                })

        ex.items.append(Item(
            item_type="template",
            title=f"{ctx.slug} calculator",
            payload={
                "template_kind": "calculator",
                "engine": "xlsx",
                "slug": ctx.slug,
                "parameters": parameters,
            },
            content_status="real",
        ))

        chain = self._calculation_chain(ctx.slug, wb_f)
        if chain:
            ex.items.append(Item(
                item_type="guidance",
                title=f"{ctx.slug} calculation chain",
                statement=chain,
                payload={"body_md": chain, "figure_ids": [], "legend_tokens": [], "disclaimer": None},
                content_status="real",
            ))

        ex.stats = {
            "sheets": len(wb_f.worksheets),
            "parameters": len(parameters),
            "inputs": sum(1 for p in parameters if p["is_input"]),
            "outputs": sum(1 for p in parameters if p["is_output"]),
        }
        return ex

    @staticmethod
    def _summarise_row(ws: Worksheet, row: int, grid_cols: set[str]) -> str | None:
        values = []
        for col_letter in grid_cols:
            v = ws.cell(row=row, column=column_index_from_string(col_letter)).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values.append(v)
        if not values:
            return None
        if len(set(values)) == 1:
            return str(values[0])
        return f"varies {min(values):g}-{max(values):g} across the period"

    @staticmethod
    def _default_value(cell, ws_v) -> str | None:
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            # formula cell: prefer the cached computed value from the data_only load
            if ws_v is not None:
                cached = ws_v[cell.coordinate].value
                if isinstance(cached, datetime.datetime):
                    return cached.date().isoformat()
                if cached is not None:
                    return str(cached)
            return None
        if isinstance(v, datetime.datetime):
            return v.date().isoformat()
        return str(v)

    @staticmethod
    def _inline_constants(ws: Worksheet) -> list[tuple[str, str]]:
        """Constants embedded directly in a formula's text rather than living
        in their own referenced cell: the 173.33 hours-per-month figure
        multiplied straight into every grade's cost formula, and the forecast
        work probability weightings (0.8 / 0.5 / 0.3 / 0.1) multiplied into
        each 'Factored at N%' row of likely (unbooked) work."""
        found: dict[str, str] = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                if "173.33" in v:
                    found.setdefault("hours_per_month", "173.33")
                m = re.search(r"\*(0\.\d+)\b", v)
                if m:
                    label = _nearest_label(ws, cell.row, cell.column_letter)
                    lm = re.search(r"factored at (\d+)%", label or "", re.I)
                    if lm:
                        found.setdefault(f"probability_weighting_{lm.group(1)}pct", m.group(1))
        return list(found.items())

    @staticmethod
    def _calculation_chain(slug: str, wb) -> str | None:
        sheet_titles = {ws.title for ws in wb.worksheets}
        if "Fee calculation" in sheet_titles:
            return (
                "Fee calculation chain: staff cost (grade hourly rate x FTE allocation x "
                "173.33 hours/month) -> + direct expenses (% of cost) -> + contingency "
                "(% of cost) -> total cost before inflation -> + inflation (compounded "
                "annually from a base rate) -> final cost -> + mark-up (% of final cost) "
                "-> fee to bill. Monthly columns are grouped under old-style RIBA work "
                "stage lettering (e.g. Stage C, Stage D&E, Stage FGH, Stage JKL). Profit "
                "percentage is read back as (minimum monthly invoice - final cost) / "
                "minimum monthly invoice."
            )
        if "Cost rates" in sheet_titles:
            return (
                "Budget and cost-rate chain: per-grade salary cost is split by the % of "
                "time charged to contracts into salary recovered on contracts vs salary "
                "left in overhead; overhead percentage = total overhead / total salary "
                "recovered on contracts; each grade's fully-burdened hourly rate = hourly "
                "cost + (hourly cost x overhead percentage). Forecast fees weight booked "
                "and prospective work by likelihood (80% / 50% / 30% / 10% factors) to "
                "build total turnover, which feeds the overall budget's gross and net "
                "profit."
            )
        if "Budget cashflow" in sheet_titles:
            return (
                "Cashflow chain: cash in = booked project income + likely (probability-"
                "weighted) work income + VAT (20% of booked income) + bank interest + "
                "other income, brought forward month to month against cash out, to give "
                "a running bank balance."
            )
        return None


def _slugify(text: str | None) -> str | None:
    if not text:
        return None
    s = re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_")
    return s or None


SPREADSHEET_CALCULATOR = SpreadsheetCalculatorExtractor()
register(SPREADSHEET_CALCULATOR)
